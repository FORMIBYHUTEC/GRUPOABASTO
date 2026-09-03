#!/usr/bin/env python3
"""Generate the idempotent Supabase catalog migration from the provided XLSX file.

Usage:
  python3 scripts/generate-catalog-sql.py \
    --workbook "/path/to/Catalogo con Precios 1.xlsx"
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

from catalog_excel_images import PUBLIC_IMAGE_PREFIX, cell_image_media_by_row


DEFAULT_OUTPUT = Path("supabase/migrations/20260831_import_catalogo_con_precios_1.sql")

# Link a spreadsheet row to a legacy product only when it is the same product
# or a clearly named equivalent, with a one-to-one relationship. Linked
# products retain their current public name, description, category and image;
# the Excel supplies the SKU, price and indicators. Every other Excel row is
# inserted as a new product with no image, ready for upload in the CMS. At the
# end, products outside the Excel catalog are removed so the table contains
# only this catalog.
LEGACY_PRODUCT_MATCHES = {
    "LIM-CLR-REPUESTO": "Cloro",
    "CLR-PERRO-REPUESTO": "Cloro de Perro",
    "CLR-PAST-KG-LMB": "Pastillas Cloro",
    "LIM-PINRP-REOUESTO": "Limpiador tipo Pinol con Repelente",
    "FAB-LL-REPUESTO": "Limpiador Multiusos Lima Limón",
    "FAB-MRFR-REPUESTO": "Limpiador Multiusos Mar Fresco",
    "FAB-CAS-REPUESTO": "Fabuloso Cascada",
    "FAB-CAMA-RESPUESTO": "Limpiador Multiusos Canela Manzana",
    "DES-IND-RESPUESTO": "Desengrasante Industrial",
    "QUI-COCH-REPUESTO": "Quitacochambre en Gel",
    "WIN-TIP-REPUESTO": "Limpiavidrios",
    "JPM-ALM-RESPUESTO": "Jabón para Manos",
    "SHOO-AUT-ESP-REPUESTO": "Shampoo para Autos",
    "SUA-MM-REPUESTO": "Suavizante",
    "SUA-DOW-REPUESTO": "Suavizante Azul",
    "JPT-DET-REPUESTO": "Jabón para Trastes",
    "VIN-AUT-REPUESTO": "Vinil para Autos",
    "DET+CLR-RESPUESTO": "Detergente para Ropa",
    "DET+NGR-RESPUESTO": "Detergente para Ropa Negra",
    "SAN-DES-REPUESTO": "Desinfectante",
    "QTS-GEL-REPUESTO": "Quitasarro en Gel",
    "QTS-LQD-REPUESTO": "Quitasarro Líquido",
    "LIM-BIC-KG": "Bicarbonato de Sodio",
    "BAÑ-PAP-GC": "Papel de Baño Rollo 180m GCPaper",
    "TOA-ROL-MRL": "Toalla de Papel Rollo 180 metros Marli",
    "SER-TOA-KIA": "Servitoalla Kihara 160",
    "ESC-CEP-CH": "Escoba Tipo Cepillo Chico",
    "ESC-4-HIL-IND": "Escoba de 4 Hilos Tipo Bruja",
    "ESC-ABA-CEP": "Escoba Cepillo de Abanico",
    "ESC-RUDO-CEP": "Escoba Uso Rudo de Cepillo",
    "CAS-CEP-NUM": "Cepillo de Castor",
    "MECH-HIL-24-IND": "Mechudo de Hilaza #24 Industrial",
    "MECH-MICR-TEL": "Mechudo de Microfibra Española",
    "MECH-MICR-TRE": "Mechudo de Microfibra Trenzada",
    "REC-CHI-PLAS": "Recogedor de Plástico Chico",
    "EXT-1.5": "Extensión de Escoba 1.5 mts",
    "TRAP-MRF-AZUL": "Trapo Microfibra Azul Membersmark",
    "CUB-NOR-10L": "Cubeta Forma Normal 10 Litros",
    "PAST-WIESE": "Pastillas Wise",
    "ATOM-1-ETI": "Atomizador Plástico",
    "BOT-5L-E": "Bote 5L",
    "BOTE-20L-BDN": "Bidón 20L",
    "BPC-812-PLY": "Bolsa de Plástico Cortada 8x12 (Salsera)",
    "ROL-POL-1525": "Bolsa en Rollo 15x25 Poliseda",
    "ROL-POL-2030": "Bolsa en Rollo 20x30 Poliseda",
    "ROL-POL-2535": "Bolsa en Rollo 25x35 Poliseda",
    "ROL-POL-3040": "Bolsa en Rollo 30x40 Poliseda",
    "ROL-PLA-1525": "Bolsa en Rollo 15x25 Plástico",
    "ROL-PLA-2030": "Bolsa en Rollo 20x30",
    "ROL-PLA-2535": "Bolsa en Rollo 25x35",
    "ROL-PLA-3040": "Bolsa en Rollo 30x40",
    "ROL-LAM-2030": "Bolsa en Rollo de Lámina 20x30",
    "BCA-CCH-2540": "Bolsa de Plástico con Asa",
    "BCA-NJU-4590": "Bolsa Negra Jumbo",
    "BBA-6090": "Bolsa Polipel 60x90",
    "BBA-90120": "Bolsa Polipel 90x120",
    "EMP-GRA-ALI": "Rollo de Emplaye 1° Grado Alimenticio",
    "EMP-18-C60": "Rollo de Emplaye 18 pulgadas Calibre 60",
    "EMQ-FLJ-NGR": "Fleje Negro",
    "FJRA-PNZ-MAN-TRSK": "Flejadora Manual Traspack con Pinzas",
    "SEG-GUA-CAR": "Guantes de Seguridad de Carnaza",
    "DES-SER-500PZ-PER": "Servilletas Paquete 500 pzas Colibrí Persa",
    "DES-SER-450PZ": "Servilletas Paquete 450 Marli Abejita",
    "DES-SER-250M": "Servilletas 1x1 Marli 250",
    "DES-CUCH-1K": "Cucharas",
    "DES-TEN-1K": "Tenedores",
    "DES-POP-KG": "Popote Desechable Económico",
    "MAN-AJE-RYB": "Papel Mantel Rojo y Blanco 1000 piezas",
    "ALU-GRU-400": "Aluminio 400",
    "CHA-007-REY": "Charola Chica",
    "CHA-CHI-066-REY": "Charola 066",
    "DES-CHA-TAQ": "Charola Chica Reyma",
    "CHA-MEGA-1014": "Charola de Unicel Muy Grande PQT 10 pzas",
    "DES-CHA-ELO": "Charola Elotera",
    "DES-CON-HAM": "Charola Hamburguesa",
    "DES-CON-77L": "Contenedor 7x7",
    "DES-CON-88L": "Contenedor 8x8",
    "VAS-12OZ-CAF": "Vaso de 12 oz para Café",
    "TAP-NEG-CAF-12": "Tapa de 12 oz para Café",
    "PLA-0-VAS": "Vaso del Cero",
    "PLA-0-TAP": "Tapa del Cero",
}


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def sql_text(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_smallint_array(values: list[int]) -> str:
    if not values:
        return "ARRAY[]::SMALLINT[]"
    return "ARRAY[" + ", ".join(str(value) for value in values) + "]::SMALLINT[]"


def source_category_for_row(row_number: int) -> str:
    if 2 <= row_number <= 37:
        return "Limpieza"
    if 39 <= row_number <= 85:
        return "Jarcería"
    if 87 <= row_number <= 123:
        return "Bolsas"
    if 125 <= row_number <= 169:
        return "Empaque y desechables"
    raise ValueError(f"Unexpected catalog row {row_number}")


def app_category(source_category: str, name: str) -> str:
    normalized = clean_text(name).lower()
    paper_terms = ("papel", "servilleta", "toalla", "ticket")
    stationery_terms = ("pluma", "marcador", "marcatextos", "grapa", "engrap", "sobre", "block", "bloc", "hoja")

    if source_category == "Limpieza":
        return "limpieza-general"
    if source_category == "Bolsas":
        return "bolsas"
    if source_category == "Jarcería":
        if any(term in normalized for term in paper_terms):
            return "papel"
        if any(term in normalized for term in stationery_terms):
            return "papeleria"
        return "utensilios"
    if source_category == "Empaque y desechables":
        if "vaso" in normalized or "tapa" in normalized:
            return "vasos"
        if any(term in normalized for term in paper_terms):
            return "papel"
        return "desechables"
    raise ValueError(f"Unexpected source category {source_category}")


def parse_indicators(value: object) -> list[int]:
    codes = [int(code) for code in re.findall(r"\d+", str(value or ""))]
    if any(code < 1 or code > 5 for code in codes):
        raise ValueError(f"Unexpected indicator value: {value!r}")
    return sorted(set(codes))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    prices_sheet = workbook["Hoja1"]
    indicators_sheet = workbook["Hoja2"]
    cell_images_by_row = cell_image_media_by_row(args.workbook)

    prices: dict[str, float | None] = {}
    for row in prices_sheet.iter_rows(min_row=2, values_only=True):
        sku = clean_text(row[0])
        if not sku or sku.startswith("Los envases de"):
            continue
        price = row[2]
        prices[sku] = float(price) if isinstance(price, (int, float)) else None

    products: list[dict[str, object]] = []
    seen_skus: set[str] = set()
    duplicate_skus: list[str] = []
    for row_number, row in enumerate(indicators_sheet.iter_rows(min_row=2, values_only=True), start=2):
        sku = clean_text(row[0])
        if not sku:
            continue
        if sku in seen_skus:
            # The source repeats PLA-4060-COR with identical data. SKU is the
            # catalog identifier, so importing it once is the only safe option.
            duplicate_skus.append(sku)
            continue
        seen_skus.add(sku)
        name = clean_text(row[1])
        source_category = source_category_for_row(row_number)
        products.append(
            {
                "sku": sku,
                "name": name,
                "category": app_category(source_category, name),
                "catalog_category": source_category,
                "indicator_codes": parse_indicators(row[4]),
                "price": prices.get(sku),
                "image_url": (
                    f"{PUBLIC_IMAGE_PREFIX}/{cell_images_by_row[row_number]}"
                    if row_number in cell_images_by_row
                    else None
                ),
            }
        )

    if len(products) != 164:
        raise ValueError(f"Expected 164 unique catalog products; found {len(products)}")
    if len(set(LEGACY_PRODUCT_MATCHES.values())) != len(LEGACY_PRODUCT_MATCHES):
        raise ValueError("Each legacy product can only be matched to one catalog SKU")
    product_skus = {product["sku"] for product in products}
    if not set(LEGACY_PRODUCT_MATCHES).issubset(product_skus):
        raise ValueError("A legacy match references an SKU that is not in the workbook")

    values = []
    sku_values = []
    for product in products:
        price = product["price"]
        price_sql = f"{price:.2f}" if isinstance(price, float) else "NULL"
        description = f"Presentación según catálogo: {product['name']}."
        values.append(
            "(" + ", ".join(
                [
                    sql_text(product["sku"]),
                    sql_text(product["name"]),
                    sql_text(description),
                    sql_text(product["category"]),
                    sql_text(product["catalog_category"]),
                    sql_smallint_array(product["indicator_codes"]),
                    price_sql,
                    sql_text(product["image_url"]),
                    sql_text(LEGACY_PRODUCT_MATCHES.get(product["sku"])),
                ]
            ) + ")"
        )
        sku_values.append("(" + sql_text(product["sku"]) + ")")

    sql = """-- ================================================================
-- Grupo Abasto | Importación de Catalogo con Precios 1.xlsx
-- Generado por scripts/generate-catalog-sql.py. No editar a mano.
-- Seguro para reejecutar: enlaza productos equivalentes sin cambiar sus
-- nombre, descripción, categoría ni imagen; después agrega los nuevos sin
-- imagen y elimina los productos fuera de este catálogo.
-- El archivo de origen repite PLA-4060-COR; se importa una única vez por SKU.
-- ================================================================

BEGIN;

ALTER TABLE products
  ADD COLUMN IF NOT EXISTS sku TEXT,
  ADD COLUMN IF NOT EXISTS catalog_category TEXT,
  ADD COLUMN IF NOT EXISTS indicator_codes SMALLINT[] NOT NULL DEFAULT ARRAY[]::SMALLINT[];

CREATE UNIQUE INDEX IF NOT EXISTS products_sku_unique
  ON products (sku)
  WHERE sku IS NOT NULL;

CREATE INDEX IF NOT EXISTS products_indicator_codes_idx
  ON products USING GIN (indicator_codes);

CREATE TABLE IF NOT EXISTS catalog_packaging_prices (
  id TEXT PRIMARY KEY,
  label TEXT NOT NULL,
  volume_liters NUMERIC(6,2) NOT NULL,
  price NUMERIC(10,2) NOT NULL CHECK (price >= 0),
  currency TEXT NOT NULL DEFAULT 'MXN',
  updated_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
);

ALTER TABLE catalog_packaging_prices ENABLE ROW LEVEL SECURITY;

INSERT INTO catalog_packaging_prices (id, label, volume_liters, price, currency)
VALUES
  ('envase-1l', 'Envase de 1 L', 1, 5.00, 'MXN'),
  ('envase-5l', 'Envase de 5 L', 5, 15.00, 'MXN'),
  ('envase-20l', 'Envase de 20 L', 20, 45.00, 'MXN')
ON CONFLICT (id) DO UPDATE SET
  label = EXCLUDED.label,
  volume_liters = EXCLUDED.volume_liters,
  price = EXCLUDED.price,
  currency = EXCLUDED.currency,
  updated_at = NOW();

WITH catalog (
  sku,
  name,
  description,
  category,
  catalog_category,
  indicator_codes,
  price,
  image_url,
  legacy_name
) AS (
VALUES
""" + ",\n".join(values) + """
),
linked_legacy_products AS (
  UPDATE products AS product
  SET
    sku = catalog.sku,
    catalog_category = catalog.catalog_category,
    indicator_codes = catalog.indicator_codes,
    price = COALESCE(catalog.price, product.price),
    image_url = COALESCE(NULLIF(product.image_url, ''), catalog.image_url),
    is_active = true
  FROM catalog
  WHERE catalog.legacy_name IS NOT NULL
    AND (product.name = catalog.legacy_name OR product.sku = catalog.sku)
    AND (product.sku IS NULL OR product.sku = catalog.sku)
  RETURNING catalog.sku
)
INSERT INTO products (
  sku,
  name,
  description,
  category,
  catalog_category,
  indicator_codes,
  price,
  image_url,
  is_active
)
SELECT
  catalog.sku,
  catalog.name,
  catalog.description,
  catalog.category,
  catalog.catalog_category,
  catalog.indicator_codes,
  catalog.price,
  catalog.image_url,
  true
FROM catalog
WHERE NOT EXISTS (
  SELECT 1
  FROM linked_legacy_products
  WHERE linked_legacy_products.sku = catalog.sku
)
ON CONFLICT (sku) WHERE sku IS NOT NULL DO UPDATE SET
  name = EXCLUDED.name,
  description = COALESCE(NULLIF(products.description, ''), EXCLUDED.description),
  category = EXCLUDED.category,
  catalog_category = EXCLUDED.catalog_category,
  indicator_codes = EXCLUDED.indicator_codes,
  price = COALESCE(EXCLUDED.price, products.price),
  image_url = COALESCE(NULLIF(products.image_url, ''), EXCLUDED.image_url),
  is_active = true;

WITH imported_skus (sku) AS (
VALUES
""" + ",\n".join(sku_values) + """
)
DELETE FROM products AS product
WHERE NOT EXISTS (
  SELECT 1
  FROM imported_skus
  WHERE imported_skus.sku = product.sku
);

COMMIT;
"""

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(sql, encoding="utf-8")
    suffix = f"; skipped duplicated SKU(s): {', '.join(sorted(set(duplicate_skus)))}" if duplicate_skus else ""
    print(f"Generated {args.output} with {len(products)} unique products{suffix}")


if __name__ == "__main__":
    main()
